#!/usr/bin/env python3
"""
build_seteuk_xlsx.py

완성된 세특 문단들을 학번/세특 내용/바이트수 3열짜리 엑셀로 만든다.
바이트수 열은 하드코딩된 숫자가 아니라 나이스(NEIS) 방식 수식으로 채운다:

    =(LENB(B2)-LEN(B2))*2+LEN(B2)

행마다 셀 위치(B2, B3, ...)만 바뀌고 수식 구조는 그대로다.

사용법 (스크립트 안의 STUDENTS 딕셔너리를 직접 채우거나, 이 파일을 참고해 비슷한
스크립트를 즉석에서 작성해도 된다):

    python3 build_seteuk_xlsx.py --json students.json --out 세특_초안.xlsx

students.json 형식:
    {"10111": "세특 문단...", "10118": "세특 문단...", "10119": "세특 문단..."}

저장 직후에는 openpyxl이 수식의 계산 결과(캐시된 값)를 갖고 있지 않아 바이트수 열이
비어 보일 수 있다. LibreOffice(soffice)를 쓸 수 있는 환경이라면 저장 후 headless
변환으로 재계산을 한 번 거쳐 값이 채워졌는지 확인하는 것이 좋다. 재계산 도구가 없는
환경이라면, 교사가 실제 엑셀에서 파일을 열 때 수식이 계산된다는 점을 안내한다.
"""

import argparse
import json

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build(students: dict, out_path: str):
    """students: {학번_또는_이름: 세특_텍스트} 순서를 유지해 전달한다 (Python dict는 삽입 순서를 보존함)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "세특_초안"

    headers = ["학번", "세특 내용", "바이트수"]
    widths = [10, 90, 12]

    font_header = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fill_header = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    font_body = Font(name="Arial", size=10)
    align_wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (title, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w

    for r, (student_id, text) in enumerate(students.items(), start=2):
        ws.cell(row=r, column=1, value=student_id).font = font_body
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=r, column=1).border = border

        b_cell = ws.cell(row=r, column=2, value=text)
        b_cell.font = font_body
        b_cell.alignment = align_wrap
        b_cell.border = border

        b_ref = f"B{r}"
        c_cell = ws.cell(row=r, column=3, value=f"=(LENB({b_ref})-LEN({b_ref}))*2+LEN({b_ref})")
        c_cell.font = font_body
        c_cell.alignment = Alignment(horizontal="center", vertical="top")
        c_cell.border = border

        ws.row_dimensions[r].height = 260

    ws.freeze_panes = "A2"

    notes = wb.create_sheet("안내")
    notes["A1"] = "바이트수 계산 안내"
    notes["A1"].font = Font(bold=True, size=12)
    lines = [
        "",
        "C열 바이트수는 =(LENB(해당 세특 셀)-LEN(해당 세특 셀))*2+LEN(해당 세특 셀) 수식으로 계산됩니다.",
        "이 수식은 나이스(NEIS) 세특 글자 수 제한 확인 시 교사들이 널리 쓰는 방식입니다.",
        "세특 내용을 수정하면 C열 값이 자동으로 다시 계산됩니다.",
    ]
    for i, line in enumerate(lines, start=2):
        notes.cell(row=i, column=1, value=line).font = Font(name="Arial", size=10)
    notes.column_dimensions["A"].width = 110

    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="세특 초안을 학번/세특 내용/바이트수 엑셀로 생성")
    parser.add_argument("--json", required=True, help="{학번: 세특텍스트} 형태의 JSON 파일 경로")
    parser.add_argument("--out", default="세특_초안.xlsx", help="출력 엑셀 파일 경로")
    args = parser.parse_args()

    with open(args.json, encoding="utf-8") as f:
        students = json.load(f)

    out_path = build(students, args.out)
    print(f"saved: {out_path}")


if __name__ == "__main__":
    main()
