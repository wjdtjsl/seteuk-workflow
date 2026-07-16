#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
batch_check.py

학생별 세특이 담긴 엑셀을 입력받아, 기계적으로 판별 가능한 항목(바이트 수, 기재금지
키워드, 관찰자 시점 의심 문장, 서식 위반)을 자동으로 표시한 1차 스크리닝 엑셀을 만든다.
개별화/구체성 제안과 humanizer 마무리, 애매한 판단(기관명인지 일반명사인지 등)은 이
스크립트가 아니라 사람(Claude + 사용자)이 출력 파일을 보고 직접 채워야 한다.

사용법:
    python3 batch_check.py <입력.xlsx> <출력.xlsx> [--name-col 이름] [--text-col 세특] [--byte-limit 1500]

입력 파일은 최소 이름(또는 학번) 컬럼과 세특(또는 내용, 특기사항) 컬럼이 있다고
가정한다. 실제 컬럼명이 다르면 --name-col / --text-col 옵션으로 지정하거나, 실행 전
Read 도구로 입력 파일 구조를 먼저 확인한다.
"""

import argparse
import re
import sys

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl --break-system-packages", file=sys.stderr)
    raise


PROHIBITED_KEYWORDS = {
    "논문": ["논문", "소논문", "학회지", "저널에 발표", "학술지 게재"],
    "도서출간": ["출간함", "저서를 발간", "책을 출판", "단행본을 출간"],
    "지식재산권": ["특허 출원", "특허 등록", "실용신안", "지식재산권"],
    "인증시험": ["토익", "토플", "텝스", "오픽", "opic", "토익스피킹",
                 "한자능력검정", "컴활", "인증시험", "어학인증"],
    "모의고사 성적": ["모의고사 등급", "모의고사 점수", "전국연합학력평가", "모의고사 성적"],
    "교외 대회 실적": ["대회 참가", "대회 수상", "상 수상", "전국대회", "교육청 주최 대회",
                    "경진대회", "공모전 수상"],
    "해외활동": ["해외연수", "해외봉사", "교환학생", "어학연수"],
    "부모/가족 배경": ["아버지가", "어머니가", "부모님 직업", "집안 형편", "가정 형편"],
    "장학 관련": ["장학생", "장학금"],
    "특정 기관/상호명": ["학원", "메가스터디", "이투스", "스카이에듀", "유네스코",
                     "구글", "네이버", "유튜브", "삼성", "카카오톡"],
    "자격증": ["자격증 취득", "자격증을 취득", "인증 획득"],
}

OBSERVER_RISK_WORDS = [
    "관심이 많음", "열정적임", "열정을 보임", "재능을 보임", "역량이 있음",
    "소질이 있음", "느낌", "깨달음", "자부심", "잠재력이 있음", "흥미를 느낌",
]

VERB_ENDING_PATTERNS = re.compile(r"(했다|한다|합니다|했음|하였다|하였음|이다)\.?\s*$")
SUBJECT_PATTERNS = re.compile(r"^(학생|본인|[가-힣]{2,4})(은|는|이|가)\s")
MIDDLE_DOT = "·"
TILDE_DATE = re.compile(r"\d{4}[.]\s*\d{2}[.]\s*\d{2}[.]?\s*~\s*\d{4}")
BOOK_HINT_WORDS = ["독후감", "책을 읽고", "도서를 읽고", "『", "책 제목"]


def split_sentences(text):
    parts = re.split(r"(?<=[.!?])\s*\n|(?<=[.!?])\s+", text)
    return [p.strip() for p in parts if p.strip()]


def count_neis_bytes(text):
    total = 0
    for ch in text:
        if "가" <= ch <= "힣":
            total += 3
        else:
            total += 1
    return total


def detect_prohibited(text):
    found = []
    for category, keywords in PROHIBITED_KEYWORDS.items():
        for kw in keywords:
            if kw in text:
                found.append("[{}] '{}'".format(category, kw))
    return found


def detect_format_issues(text):
    issues = []
    sentences = split_sentences(text)
    for s in sentences:
        if VERB_ENDING_PATTERNS.search(s):
            issues.append("명사형 종결 아님: \"{}\"".format(s[-20:]))
        m = SUBJECT_PATTERNS.match(s)
        if m:
            issues.append("문두 주어 사용: \"{}\" in \"{}...\"".format(m.group(0).strip(), s[:20]))
    if MIDDLE_DOT in text:
        issues.append("특수기호(가운데점) {}회 사용".format(text.count(MIDDLE_DOT)))
    if TILDE_DATE.search(text):
        issues.append("날짜에 물결표(~) 사용 의심 (하이픈 형식 0000.00.00.-0000.00.00. 확인 필요)")
    for hint in BOOK_HINT_WORDS:
        if hint in text:
            issues.append("도서 관련 서술 감지: 단순 독후감이면 도서명 입력 불가 규정 확인 필요")
            break
    return issues


def detect_observer_risk(text):
    risky = []
    for word in OBSERVER_RISK_WORDS:
        if word in text:
            for s in split_sentences(text):
                if word in s:
                    risky.append("'{}' 포함: \"{}\"".format(word, s))
    return risky


def find_column(header, candidates):
    for idx, col in enumerate(header):
        if col is None:
            continue
        col_str = str(col).strip()
        for cand in candidates:
            if cand in col_str:
                return idx
    return None


def main():
    parser = argparse.ArgumentParser(description="세특 엑셀 일괄 1차 스크리닝")
    parser.add_argument("input_xlsx", help="입력 파일 (학생별 세특 원본)")
    parser.add_argument("output_xlsx", help="출력 파일 (검토 결과)")
    parser.add_argument("--name-col", default=None, help="이름/학번 컬럼명")
    parser.add_argument("--text-col", default=None, help="세특 내용 컬럼명")
    parser.add_argument("--byte-limit", type=int, default=1500, help="NEIS 바이트 제한")
    args = parser.parse_args()

    wb = load_workbook(args.input_xlsx)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("입력 파일에 데이터가 없습니다.", file=sys.stderr)
        sys.exit(1)

    header = list(rows[0])

    if args.name_col:
        name_idx = find_column(header, [args.name_col])
    else:
        name_idx = find_column(header, ["이름", "성명", "학번", "번호"])

    if args.text_col:
        text_idx = find_column(header, [args.text_col])
    else:
        text_idx = find_column(header, ["세특", "특기사항", "내용"])

    if text_idx is None:
        print("세특 내용 컬럼을 찾지 못했습니다. --text-col 옵션으로 정확한 컬럼명을 지정하세요.",
              file=sys.stderr)
        print("발견된 헤더: {}".format(header), file=sys.stderr)
        sys.exit(1)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "세특 검수 결과"

    out_header = [
        "이름/학번", "원본", "바이트수",
        "제한({}Byte) 초과 여부".format(args.byte_limit),
        "1. 절대 기재금지", "2. 관찰자 시점 의심", "3. 서식 규정 위반",
        "4. 개별화/구체성 제안 (직접 작성)",
        "5. 최종본 (규정 수정 + humanizer 마무리, 직접 작성)",
    ]
    out_ws.append(out_header)
    for col_idx in range(1, len(out_header) + 1):
        cell = out_ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    processed = 0
    for row in rows[1:]:
        text = row[text_idx] if text_idx < len(row) else None
        if not text or not str(text).strip():
            continue
        text = str(text).strip()
        name = row[name_idx] if (name_idx is not None and name_idx < len(row)) else ""

        byte_count = count_neis_bytes(text)
        over_limit = "초과" if byte_count > args.byte_limit else "이내"
        prohibited = detect_prohibited(text)
        format_issues = detect_format_issues(text)
        observer_risk = detect_observer_risk(text)

        out_ws.append([
            name,
            text,
            byte_count,
            over_limit,
            "\n".join(prohibited) if prohibited else "해당 없음",
            "\n".join(observer_risk) if observer_risk else "해당 없음",
            "\n".join(format_issues) if format_issues else "해당 없음",
            "",
            "",
        ])
        processed += 1

    widths = [12, 45, 10, 14, 28, 28, 28, 32, 45]
    for i, w in enumerate(widths, start=1):
        out_ws.column_dimensions[get_column_letter(i)].width = w
    for row in out_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_wb.save(args.output_xlsx)
    print("완료: {}건 처리, 결과 저장 -> {}".format(processed, args.output_xlsx))
    print("주의: 기재금지/관찰자시점/서식 항목은 키워드 기반 1차 스크리닝입니다. "
          "실제 문맥 판단, 개별화 제안, humanizer 마무리를 반영한 최종본은 사람이 직접 작성해야 합니다.")


if __name__ == "__main__":
    main()
